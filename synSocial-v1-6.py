"""
 SynSocial
 Social Media (Twitter) Data Generator
 Source point of contact - Anne Tall
 Date = 15 November 2021
 
"""

# Import libraries(Standard, Third party, Local application/library)
import sys
import os
import json
import csv
import pprint
import uuid
import random
import logging
import time
from random import seed
from random import randint
import datetime
from datetime import datetime, timedelta
from datetime import date

import xlsxwriter
import xlrd
import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
import numpy as np
import codecs

# Define constants
TED = datetime(2007, 1, 1, 00, 00, 00)  # Twitter Establishment
# Date = year=2007, month=1=January, date= 1

TODAY = datetime(2022, 1, 1, 00, 00, 00)  # Date of last Twitter message (Tweet),
# if the patient's death date in the input file is later or not present
# Date = year=2020, month=1=January, date= 1 

# Paths and reference files
# For Windows - format is ".\\folder_name" for data file and ".\folder_name for output"
# For Linux - format is "./folder_name" for all folders
localDataFilePath = "./data"  # File path for Synthea FHIR JSON data files that are input to the program
localOutputFilePath = "./output"
localRefDataFilePath = "./ref-data"  # Folder

# that contains the reference files (rate-info, content, baseline, nounlist, phrases)
excelOutputFileName = "output-stats.xlsx"
excelRefDataFileName = "rate-info.xls"  # Reference data Excel Workbook file name - name

# of the file that contains the rate, impact to baseline message generation for each code
excelRefCodeWks = "Code-Ref-Rates"  # The Worksheet in the rate-info Excel Workbook
# with condition code rates of Twitter message generation
excelRefBaseWks = "Baseline-Ref-Rates"  # The Worksheet in the rate-info Excel Workbook
# with baseline rates of Twitter message generation

"""
Replaced Excel Spreadsheets with message text with Text Files
excelCodeContentFileName = "content.xls"  # The file that contains the messages contents
# for each condition code Twitter message
excelBaseContentFileName = "baseline.xls"  # The file that contains the messages contents
# for each baseline Twitter message
"""
localTextMessageFilePath = "/home/anne/twitter-scraper/output"
#see comment above for format of folder name


randomNounsFileName = "nounlist.txt"  # The file that contains words used to create the
# Twitter screen name
fileNamePhrases = "phrases.xls"  # The file that contains phrases added to the
# beginning and end of the Tweet text
excelOutputSheetBaseline = "Daily Baseline Tweet Stats" #Excel Worksheet in output-stats for the
# Daily Baseline Tweet Statistics
excelOutputSheetCondition = "Daily Condition Tweet Stats" #Excel Worksheet in output-stats for the
# Daily Condition Tweet Statistics



def config_stats_sheet(input_filenames_processed):
    """
     Creates the Excel Spreadsheet row and column lables for recording the number of messages generated
     Used in first line of main
     Accepts an age in days as input and returns the Baseline mean Tweet generation rate and 
     name of the range
    """
    
    #Write Sheet Names
    outputStatsWorkbook = xlsxwriter.Workbook("{}/{}".format(localOutputFilePath,
                                                                         excelOutputFileName))
    outputStatsBaselineSheet = outputStatsWorkbook.add_worksheet(excelOutputSheetBaseline)
    outputStatsConditionSheet = outputStatsWorkbook.add_worksheet(excelOutputSheetCondition)

    #Write Column Titles
    outputStatsBaselineSheet.write(1, 0, "Day")
    outputStatsConditionSheet.write(1, 0, "Day")

    outputStatsBaselineSheet.write(0, 1, "# of Baseline Tweets")
    outputStatsConditionSheet.write(0, 1, "# of Condition Tweets")


    patient_short_filename=[]
    patient_short_filename_list=[]

    for item in input_filenames_processed:
        patient_filename = item ['fileName']
        patient_short_filename = patient_filename [0:9]
        patient_short_filename_list = [patient_short_filename] + [*patient_short_filename_list]

    n = len (patient_short_filename_list)
    row= 1
    col = 1
    for z in range (0, n):
        outputStatsBaselineSheet.write(row, col, patient_short_filename_list[z])
        outputStatsConditionSheet.write(row, col, patient_short_filename_list[z])
        col=col+1

    #Write Row Dates
    dayTED = TED
    dayTODAY = TODAY

    row = 2
    col = 0
    dd = [dayTED + timedelta(days=x) for x in range ((dayTODAY-dayTED).days+1)]
    for y in dd:
        outputStatsBaselineSheet.write(row, col, y.strftime("%m/%d/%Y"))
        outputStatsConditionSheet.write(row, col, y.strftime("%m/%d/%Y"))
        row=row+1

    outputStatsWorkbook.close()
    return 

def record_stats_day (input_day, num_generation, input_type, input_index):
    """
    Writes the number of baseline and medical condition code Tweets created on that day
    Used by generate_tweets function
    accepts as input the day, filename (for the user / patient ), the number
     of Tweets generated that day, and the type of messsage (baseline or code)
    Writes the statistics to the Output Excel file and returns
    """

    if (input_type == 'baseline'):
        outputStatusWorksheet = outputStatusWorkbook[excelOutputSheetBaseline]
    elif (input_type == 'code'):
        outputStatusWorksheet = outputStatusWorkbook[excelOutputSheetCondition]

    inputDate = datetime(input_day.year, input_day.month, input_day.day, 0, 0, 0)
    i = (inputDate - TED).days + 3
    j = input_index

    cellCurrent = outputStatusWorksheet.cell(row = i, column = j ).value 
    if (cellCurrent is None):  
        cellCurrent = 0
    cellFilled = outputStatusWorksheet.cell(row = i, column = j)
    cellFilled.value = num_generation[0] + cellCurrent

    return

def obtain_mean_base_rate(input_age):
    """
     Provides the Baseline Twitter Message Mean generation Rate Based upon Age Group 
     Used in codes_in_effect function
     Accepts an age in days as input and returns the Baseline mean Tweet generation rate and 
     name of the range
    """

    base_mean_out = 0
    base_range_out = "under18"

    lookup_age = int(input_age / 365)

    if lookup_age < 18:
        base_mean_out = 0
        base_range_out = "under18"
        return base_range_out, base_mean_out
    elif lookup_age < 20:
        base_range_out = "18to20"
    elif lookup_age < 25:
        base_range_out = "20to25"
    elif lookup_age < 30:
        base_range_out = "25to30"
    elif lookup_age < 40:
        base_range_out = "30to40"
    elif lookup_age < 50:
        base_range_out = "40to50"
    elif lookup_age < 60:
        base_range_out = "50to60"
    elif lookup_age < 70:
        base_range_out = "60to70"
    elif lookup_age < 85:
        base_range_out = "70to85"
    else:
        base_range_out = "85to99"

    base_mean_out = [base_reference_workbook["meanRate"][i] for i in range (len(base_reference_workbook))\
                    if base_reference_workbook["Age"][i] == base_range_out]

    return base_range_out, base_mean_out


def obtain_mean_code_rate(input_condition_code):
    """
     Provides the Condition Twitter Message Mean generation Rate Based upon Medical Condition Code 
     Used in codes_in_effect function
     Accepts a Condition Code as input and returns the Condition mean Tweet generation rate, 
      and impact on the Baseline Rate (alpha)
    """
    for i in range (len(code_reference_workbook)):
      if str(code_reference_workbook["code"][i]) == str(input_condition_code):
          mean_out = code_reference_workbook["meanRate"][i]
          alpha_out = code_reference_workbook["alpha"][i]
 
    return mean_out, alpha_out


def compute_twit_end_date(input_deceased_date, input_birth_date):
    """
     Determines the End Date for Tweet Generation 
     Used in obtain_twitter_dates function
     Accepts deceased date and birth date as input and returns
      Twitter Message (Tweet) generation end date
    """

    age_at_death = input_deceased_date - input_birth_date
    if input_deceased_date == "" or input_deceased_date == datetime(9999, 1, 1, 00, 00, 00):
        twitter_enddate_out = TODAY
    elif input_deceased_date > TODAY:
        #print("error:  died in the future, update date for TODAY")
        twitter_enddate_out = 0
    elif age_at_death > timedelta(days=36135):  # 36135 days in 99 years
        twitter_enddate_out = input_birth_date + relativedelta(years=+99)
    else:
        twitter_enddate_out = input_deceased_date
    return twitter_enddate_out


def test_abate_date(input_abate_date, input_twit_start_date):
    """
     Tests the Medical Condition Code Abatement Date 
     Used in obtain_tweet_conditions function
     Returns TRUE if abatementDateTime > twitStartDate 
     The Medical condition does not apply if it ends before start date for sending Tweets
    """

    if input_abate_date > input_twit_start_date:
        return True
    else:
        return False


def test_onset_date(input_onset_date, input_twitter_end_date):
    """
     Tests the Medical Condition Code OnSet Date 
     Used in obtain_tweet_conditions function
     Returns TRUE if onsetDateTime < twitEndDate 
     The Medical condition code does not apply if it started after Tweet end date
    """

    if input_onset_date < input_twitter_end_date:
        return True
    else:
        return False


def codes_in_effect(input_list_codes, input_age):
    """
     Combines the message generation rates for Medical Conditions 
      and the Baseline Age-Group messages
     Used in the generate_tweets function
     Provides the combined mean Twitter message generation rate and the list of 
      effective codes for the particular date using the list of codes input and 
      age at that date 
    """

    alpha_list = [1]
    effective_codes = []
    num_codes = len(input_list_codes)
    for i in range(0, num_codes):
        check_code = input_list_codes[i]
        mean, alpha = obtain_mean_code_rate(check_code)  # Call the read code rates function

        code_dict = {
            'code': check_code,
            'meanvalue': mean,
            'type': 'code'
        }
        effective_codes.append(code_dict)

        alpha_list.append(float(alpha))  # the impact (alpha) to the baseline message generation rate is 
                                         # the minimum alpha value assigned to all codes

    baseline_name, base_mean_raw = obtain_mean_base_rate(input_age)  # Call baseline rate function
    alpha_min = min(alpha_list)
    base_mean = float(base_mean_raw[0]) * alpha_min
    code_b_dict = {
        'code': baseline_name,
        'meanvalue': base_mean,
        'type': 'baseline'
    }
    effective_codes.append(code_b_dict)
    return effective_codes


def gen_screen_name(input_name):
    """
     Creates Twitter User Screen Name
     Used in obtain_twitter_dates function
     The screen name is generated by combining the Patient Name and a random word 
     For a particular Patient, the Username is the same for all Tweets
     Each time the program is run a new Twitter Screen Name is generated for a Patient
    """
    x = random.random()
    random_sel = int(num_nouns *x) #randint(0, num_nouns)
    noun = nounList[random_sel]
    tweet_screenname_output = "@" + input_name[0:6] + noun[0:7]
    return tweet_screenname_output

def gen_user_id():
    """
     Creates Twitter User Id 
     Used in obtain_twitter_dates function
     Based upon current date-time (random sequential value) 
     Each time the program is run a new Twitter User Id is generated for a Patient
    """

    now = datetime.now()  # 20 digits
    format_now_day_time = "%y%m%d%H%M%S%f"  # 18 digits
    day_time_string = now.strftime(format_now_day_time)
    user_id_out = str(day_time_string)
    return user_id_out


def output_tweet_json(
        input_day, input_code, input_type, input_name,
        input_screenname, input_user_id, input_location, input_file_name, input_first_msg_bool):
    """
     Writes to output file Tweets in Twitter JSON Format
     Used in generate_tweets function
     Tweets are Appended to a file nammed to correspond to the
      Synthea JSON file name 
     Tweet Text are limited to 140 characters, 
     (Extendded 280-character Tweets are not generated) 
     Creates original Tweet posts, and does not include ReTweets, Likes and Comments
    """

    created_at = gen_twit_date_time(input_day)
    id_str_gen = gen_twit_id_str(created_at)
    tweet_text = obtain_tweet_text(input_code, input_type)
    user_id_gen = int(input_user_id)
    user_id_gen_str = str(input_user_id)
    name_gen = input_name
    screen_name_gen = input_screenname
    place_name = input_location
    fundemental_tweet={ }

    if (input_first_msg_bool == False):

        # Message to be appended

        fundemental_tweet = {
            'created_at': created_at,  # UTC time when the tweet was created
            # format of time is:  three characters for day of week, month, day of month
            # hour, colon, minute, colon, second, four numbers, year
            'id_str': id_str_gen,  # String representation of the unique Twitter
            # message id (which is a 64 bit integer)
            # the unique ID is based on time, worker number and seuqence number
            'text': tweet_text,  # 140 character message
            'user': {
                'id': user_id_gen,
                'id_str': user_id_gen_str,
                'name': name_gen,
                'screen_name': screen_name_gen,
                'location': None
            },  # author of the tweet
            'place': {'country': "United States", 'name': place_name},
            'entities': {'hashtags': [], 'urls': []},
            # hashtags, user mentions, URLs, cashtags, native media
            # (photos, videos, animated GIFs)
            'extended_entities': {'media': []}  # attached or native media
        }

        filename = input_file_name[ :-5] + "TwitterData" + ".json"  
        # file name = original filename with TwitterData appended on the end
        
        # Added TRY to handle errors

        try:
            with open("{}/{}".format(localOutputFilePath, filename), "r+") as outfile1:
                file_data = json.load(outfile1)         
                file_data["messages"].append(fundemental_tweet)
                outfile1.seek(0)
                json.dump(file_data, outfile1, indent=4)  ######## Revise to make a bundle
                # generated tweet appended to the file
                outfile1.close()
        except:
            with open("{}/{}".format(localOutputFilePath, filename), "w+") as outfile:
                outfile.seek(0)
                fundemental_tweet_one = {'messages':[fundemental_tweet]}
                json.dump(fundemental_tweet_one, outfile, indent=4)  
                # second generated tweet used to create the file
            outfile.close() 

    else:

    # First Message to be written

        fundemental_tweet_one = {'messages': [{
            'created_at': created_at,  # UTC time when the tweet was created
            # format of time is:  three characters for day of week, month, day of month
            # hour, colon, minute, colon, second, four numbers, year
            'id_str': id_str_gen,  # String representation of the unique Twitter
            # message id (which is a 64 bit integer)
            # the unique ID is based on time, worker number and seuqence number
            'text': tweet_text,  # 140 character message
            'user': {
                'id': user_id_gen,
                'id_str': user_id_gen_str,
                'name': name_gen,
                'screen_name': screen_name_gen,
                'location': None
            },  # author of the tweet
            'place': {'country': "United States", 'name': place_name},
            'entities': {'hashtags': [], 'urls': []},
            # hashtags, user mentions, URLs, cashtags, native media
            # (photos, videos, animated GIFs)
            'extended_entities': {'media': []}  # attached or native media
        } ] }

        
        filename = input_file_name[ :-5] + "TwitterData" + ".json"  
        # file name = original filename plus twitter data
        with open("{}/{}".format(localOutputFilePath, filename), "w+") as outfile:
            outfile.seek(0)
            json.dump(fundemental_tweet_one, outfile, indent=4)  
            # generated first tweet added to the file
            outfile.close() 
    return


def gen_twit_date_time(input_day):
    """ 
     Creates Twitter Formated Time 
     Used in output_tweet_json function
     The Time (Hour, Minutes and Seconds) are randomly generated from input Day 
    """

    #now = datetime.now()

    # randomly generate hour
    #hour_seed = now.hour
    #seed = hour_seed
    x=random.random()
    hour_value = int(23*x) #randint(0, 23)

    # randomly generate minute
    #minseed = now.minute
    #seed = minseed
    x=random.random()
    min_value = int(59*x) #randint(0, 59)

    # randomly generate second
    #secseed = now.second
    #seed = secseed
    x=random.random()
    sec_value = int(59*x) ##randint(0, 59)

    time_string = str(hour_value) + "," + str(min_value) + "," + str(sec_value)
    day_time_string = str(input_day) + " " + time_string
    format_day_time1 = "%Y-%m-%d %H,%M,%S"
    day_time_string_twitter = datetime.strptime(day_time_string, format_day_time1)
    created_day_time_out = day_time_string_twitter.strftime("%a %b %d %H:%M:%S +0000 %Y")

    return created_day_time_out


def gen_twit_id_str(input_created_at):
    """
     Creates Twitter Formated Message Id 
     Used in output_tweet_json function
     Returns the 'id_str' field 
     IDs are unique 64-bit unsigned integers, which are based on time, 
      instead of being sequential.
     The full ID is composed of a timestamp, a worker number, and a sequence number.
    """

    now = datetime.now()
    microseed = now.microsecond
    seed = microseed
    x1 = random.random()
    x2 = random.random()
    x3 = random.random()
    if x1 == 0:
        x1=0.1
    if x2 == 0:
        x2 = 0.1
    if x3 == 0:
        x3 = 0.1
    rand_ip2 = int(254*x1) #randint(1, 254)
    rand_ip3 = int(254*x2) #randint(1, 254)
    rand_ip4 = int(254*x3) #randint(1, 254)

    worker_number = "10" + str(rand_ip2) + str(rand_ip3) + str(rand_ip4)
    sequence_number = 1

    format_day_time2 = "%a %b %d %H:%M:%S %z %Y"
    day_time_input = datetime.strptime(input_created_at, format_day_time2)
    time_base = day_time_input.strftime('%w%m%d%H%M%S%Y')
    id_str_out = str(time_base) + worker_number + str(sequence_number)
    return id_str_out


def obtain_tweet_text(input_code, input_type):
    """
     Creates Text of Tweet 
     Used in output_tweet_json function.
     Creates the Tweet text based upon the effective Medical Condition codes 
      and Baseline Age-Group
     Text is randomly selected from a row in a TEXT file (previously Excel file)
     A random phrase is added to the start and end of each text 
     One file containing all the Tweets each Patient is Created
    """

    text_content_filename = input_code + ".txt" #excelCodeContentFileName
    
    #UseText Files (replacing Spreadsheet) for Tweet Text
    #randomly select a lines in the text file

    file_name_dir = os.path.join(localTextMessageFilePath, text_content_filename)

    with codecs.open (file_name_dir, "r", encoding="utf8", errors='ignore') as file:
        counter = 0

        content = file.read()
        content_lines = content.split("\n")

        text_selected = random.choice(content_lines)
    file.close()

    # randomly select start and end phrase
    # from the list in phrases.xls, start and end sheets
    
    num2 = len(df_start_textlist) - 1
    num3 = len(df_end_textlist) - 1
    x=random.random()
    sel_start_text = int(num2*x) #randint(0, num2)
    sel_end_text = int(num3*x) #randint(0, num3)
    col_p = 0
    row_s = sel_start_text
    row_e = sel_end_text
    start_text_selected = df_start_textlist.iat[row_s, col_p]
    end_text_selected = df_end_textlist.iat[row_e, col_p]

    tweet_text_out = start_text_selected + ", " + text_selected + ", " + end_text_selected
    return tweet_text_out

def read_FHIR_dates():
    """
     Read dates from the Patient Healthcare Information (JSON file) Generated by Synthea 
     Reads each Patient's Birthdate and Deceased-date
      and Creates data_dates dictionary 
    """

    data_dates = []  # Declares the list for dates that are used to

    if os.path.exists(localDataFilePath):
        # generate the Twitter data
        filename_processed = []  # Declares the list for the names of
        # the files processed

        num_files = 0
        for filename in os.listdir(localDataFilePath):

            with codecs.open("{}/{}".format(localDataFilePath, filename), "r", encoding="utf8", errors='ignore') as f:
                try:
                    data = json.load(f)

                    for i in data['entry']:  # Get patient dates
                        x = i['resource']
                        if "Patient" == x['resourceType']:
                            data_dict_one = {
                                'birthDate': x['birthDate'] if x.get('birthDate') else "0001-1-1",
                                'deceasedDateTime': x['deceasedDateTime'].split("T")[0] if x.get(
                                    'deceasedDateTime') else "9999-1-1",
                                'name': x['name'][0]['given'][0] if x.get('name') else "",
                                'city': x['address'][0]['city'] if x.get('address') else "",
                                'state': x['address'][0]['state'] if x.get('address') else "",
                                'fileName': filename
                            }
                            num_files = num_files + 1
                            filename_processed.append(filename)  # list of files processed
                            data_dates.append(data_dict_one)  # add each date to the dictionary
                except json.JSONDecodeError:  # Handling errors
                    print("JSON file does not exist")
    else :
        print("Local directory does not exist")
    f.close()
    return data_dates


def obtain_twitter_dates(data_dates):
    """
    Tests each Patient's Birth Date and Deceseased Date Test 
     
     Determines if Birthdate, Deathdates correspond to the Twitter Establishment Date (TED), 
     such that they would be valid users of Twitter
     Create data_dates_twitter for Patients that are valid users
     Create filenames_unprocessed for Patients that are NOT valid users

     For each Patient, determines if they do not create Twitter Messages (Tweets)
     Patients do not create Tweets if their deceased Date is before TED,
      if deceasedDateTime < = TED

     For each Patients that creates Twitter Messages (else - deceasedDateTime > TED)
     Determines the twitStartDate and twitEndDate - these values are added to the data_dates_twitter dictionary
     twitStartDate is the date of the first twitMessage (data type is date-time)
     twitEndDate is the date of the last twitMessage (data type is date-time)

     twitStartDate id determined by setting check_twitter_startdate = birthDate + 18 years
      either (1) the person is over the age of 18 when Twitter became avalble
      - if checkTwitSartDate >TED, then twitStartDate = TED
      or (2) the start date is when they are 18 years old - else twitStartDate = check_twitter_startdate  
    """
    
    filenames_unprocessed = []
    filenames_processed = []
    data_dates_twitter = []
    for values in data_dates:
        decease_testdate = values['deceasedDateTime']
        decease_tested_date = datetime.strptime(decease_testdate, "%Y-%m-%d")
        birth_testdate = values['birthDate']
        birth_tested_date = datetime.strptime(birth_testdate, "%Y-%m-%d")
        check_twitter_startdate = birth_tested_date + relativedelta(years=+18)  # date at 18
        if decease_tested_date <= TED or decease_tested_date <= check_twitter_startdate:  # Died before Twitter
            # Establishment Date or Died before 18th Birthday
            null_data = {'fileName': values['fileName']}
            filenames_unprocessed.append(null_data)
        elif check_twitter_startdate > TED:
            twitter_enddate_from_function = compute_twit_end_date(decease_tested_date, birth_tested_date)
            new_data = {
                'birthDate': values['birthDate'],
                'deceasedDateTime': values['deceasedDateTime'],
                'name': values['name'],
                'tweetScreenName': gen_screen_name(values['name']),
                # use the function to generate the Twitter Screen Name based upon Name
                'tweetUserIdStr': gen_user_id(),  # call function
                # to create the string of the Twitter User Id
                'city': values['city'],
                'state': values['state'],
                'fileName': values['fileName'],
                'twitStartDate': datetime.strftime(check_twitter_startdate, "%Y-%m-%d"),
                # Twitter Start Date = date turned 18 since this occurred after TED
                'twitEndDate': datetime.strftime(twitter_enddate_from_function, "%Y-%m-%d")
            }
            data_dates_twitter.append(new_data)
            file_data = {'fileName': values['fileName']}
            filenames_processed.append(file_data)
        else:
            twitter_enddate_from_function = compute_twit_end_date(decease_tested_date, birth_tested_date)
            new_data = {
                'birthDate': values['birthDate'],
                'deceasedDateTime': values['deceasedDateTime'],
                'name': values['name'],
                'tweetScreenName': gen_screen_name(values['name']),
                # use the Tweet Name Generation function to create the Twitter Screeb Name
                'tweetUserIdStr': gen_user_id(),  # call function to
                # create the string of the Twitter User Id
                'city': values['city'],
                'state': values['state'],
                'fileName': values['fileName'],
                'twitStartDate': datetime.strftime(TED, "%Y-%m-%d"),  # Twitter
                # Start Date = TED since turned 18 before TED
                'twitEndDate': datetime.strftime(twitter_enddate_from_function, "%Y-%m-%d")
            }
            data_dates_twitter.append(new_data)
            file_data = {'fileName': values['fileName']}
            filenames_processed.append(file_data)
    return data_dates_twitter, filenames_unprocessed, filenames_processed


def obtain_tweet_conditions(data_dates_twitter):
    """
     Adds the Effective Medical Condition Codes 
      to the input DataDatesTwit dictionary and returns 
     Includes codes that have abatement Dates after the Twitter Start Date and
     Includes codes that have onsetDates before the Twitter End Date
    """


    for patdata in data_dates_twitter:
        filenamepat = patdata['fileName'] if patdata.get('fileName') else ""
        condition_twitter = []  # List of DataDatesTwit with condition
        for filenamelist in os.listdir(localDataFilePath):
            if filenamepat == filenamelist:
                with codecs.open("{}/{}".format(localDataFilePath, filenamelist), "r", encoding="utf8", errors='ignore') as f:
                    print("    ", patdata['fileName'])
                    try:
                        data = json.load(f)
                        for i in data['entry']:  # for each code
                            x = i['resource']
                            if "Condition" == x['resourceType']:  # if the resource type is "Condition"
                                # add a dictionary of effective conditions
                                test_start_date = x['onsetDateTime'] if x.get('onsetDateTime') else "0001-01-01"
                                added_test_abatement_date = test_start_date
                                test_abatement_date = x["abatementDateTime"] if x.get("abatementDateTime") else "9999-01-01"
                                test1 = test_abate_date(test_abatement_date, patdata[
                                    'twitStartDate'])  # Check to confirm the condition
                                # AbatementDate is later than the twitStartDate
                                test2 = test_onset_date(test_start_date, patdata[
                                    'twitEndDate'])  # check to confirm the Onset of the
                                # condition is before the twitEndDate
                                if test1 and test2:
                                    holdOnsetDate = x['onsetDateTime'].split("T")[0] if x.get('onsetDateTime') else "0000-01-01"
                                    dateHoldOnsetDate = datetime.strptime(holdOnsetDate, "%Y-%m-%d")
                                    forced_abatement = dateHoldOnsetDate + timedelta(days=365)
                                    str_forced_abatement = forced_abatement.strftime ("%Y-%m-%d")
                                    data_dict_codes = {
                                        'code': x['code']["coding"][0]['code'] if x.get('code') else "null",
                                        'onsetDateTime': x['onsetDateTime'].split("T")[0] if x.get('onsetDateTime') else "0000-01-01",
                                        'abatementDateTime': x['abatementDateTime'].split("T")[0] if x.get('abatementDateTime') else str_forced_abatement
                                    }
                                    condition_twitter.append(
                                        data_dict_codes)  # Append dictionary of codes
                                    # to each patient dictionary in list
                    except json.JSONDecodeError:  # Handling errors
                        print("JSON file does not exist")

        patdata['conditions'] = condition_twitter
        f.close()
    return data_dates_twitter 


def generate_tweets(data_dates_twitter):
    """
     Generates Tweets based upon the Daily Message Generation Rate 
     For Each Date in range num_dates_tweet, identify the codes and base rate in effect
      for that day - effective_code_date
     
     For each date - Generate Twitter messages:
     Time on Message = twitter_startdate_formatted + i 
     Text of Message = column [1] in content.xls of worksheet with tab that matches the code and 
      in column[1] in baseline.xls with tab matching age range
     Author of message, is the Twitter user screen name based upon the patient’s name = item ['name']
     Number of messages = Poisson distrubted random number with mean returned from the function
    """
    

    first_msg = True
    name_list = []

    global outputStatusWorkbook
    outputStatusWorkbook = openpyxl.load_workbook("{}/{}".format(localOutputFilePath, excelOutputFileName))
    for item in data_dates_twitter:
        twitter_startdate_formatted = datetime.strptime(item['twitStartDate'], "%Y-%m-%d").date()
        twitter_enddate_formatted = datetime.strptime(item['twitEndDate'], "%Y-%m-%d").date()

        num_dates_tweet = (twitter_enddate_formatted - twitter_startdate_formatted).days

        birthdate_formatted = datetime.strptime(item['birthDate'], "%Y-%m-%d").date()
        age_at_start = (twitter_startdate_formatted - birthdate_formatted).days
            
        #######
        # num_dates_tweet = 50  # Override for test #######
        # remove comment on previous line to limit the number of messages generated 
        #######

        printData = ( item['name'], item['fileName'], len(item['conditions']))
        print("    ", printData)
        

        # Get Name index from excel file        
        outputStatusWorksheet = outputStatusWorkbook[excelOutputSheetBaseline]            
        lastCol = outputStatusWorksheet.max_column
        name_index = 0
        patientShortFilename = item['fileName'][0:9]
        for i in range (2, lastCol + 1):
            testcell = outputStatusWorksheet.cell (row = 2, column=i).value
            if str(testcell) == patientShortFilename :
                name_index = i
            
        for i in range(0, num_dates_tweet):
            age = age_at_start + i
            twitter_date = twitter_startdate_formatted + timedelta(days=i)

            num_condition = len(item['conditions']) 
            #if i % 10 == 0:
            #    print("    ", twitter_date, end='\r', flush=True)

            code_list = []
            for j in range(0, num_condition):
                onset_datetime_formatted = datetime.strptime(item['conditions'][j]['onsetDateTime'], "%Y-%m-%d").date()
                abatement_datetime_formatted = datetime.strptime(item['conditions'][j]["abatementDateTime"], "%Y-%m-%d").date()
                test3 = onset_datetime_formatted <= twitter_date
                test4 = abatement_datetime_formatted >= twitter_date
                if test3 and test4:
                    code_list.append(item['conditions'][j]['code'])
            
            effective_code_date = codes_in_effect(code_list, age)  # Call function to obtain codes in effect

            for code in effective_code_date:
                # Create a Twitter message for each code for the twitter_date
                input_day = twitter_date  # Date of the created Twitter Message
                input_code = code['code']
                input_type = code['type']
                input_name = item['name']
                input_screenname = item['tweetScreenName']
                input_user_id = item['tweetUserIdStr']
                input_location = item['city'] + ", " + item['state']
                input_file_name = item['fileName']

                # execute the function that generates the output the mean value number of times
                mean = code['meanvalue']

                num_generation = np.random.poisson(mean, 1)  # The number generated is 
                # Poisson Distributed from the Mean
                               
                record_stats_day(input_day, num_generation, input_type, name_index) 
                #Capture Statistic - Number of baseline and medical condition messages generated on that day for that user
                
                msg_count = num_generation
                for j in range(1, int(num_generation)):
                    #check to see if new name, new file or addition to current file
                    val = str(input_name)
                    if val in name_list:
                        first_msg = False
                    else:
                        name_list.append(val)
                        first_msg = True
                    
                    #call the function to generate and output the Tweet
                    output_tweet_json(input_day, input_code, input_type, input_name, input_screenname, input_user_id, input_location, input_file_name, first_msg)  
    
    outputStatusWorkbook.save("{}/{}".format(localOutputFilePath, excelOutputFileName))                    
    print("Results:  Finished generating files of Twitter messages")
    return 

def output_statistics(inputfilenames_processed, inputfilenames_unprocessed):
    """
    Output Statistics Associated with the Data Processed
    Output the Number of Synthea Files Input and Processed and the 
    Number of Social Media Message Sets (Sets of Tweets) Generated
    """

    numFilesTotal = len(inputfilenames_processed)+len(inputfilenames_unprocessed)
    numUnProcessed = len(inputfilenames_unprocessed)
    numFiles = len(inputfilenames_processed)

    print("The social media message generation end date used was:", TODAY.strftime("%m/%d/%Y"), " (month/day/year)")
    print("The total number of Synthea (patient JSON) files input is: ", numFilesTotal)
    print("The total number of Synthea files not processed is: ", numUnProcessed)
    print("The total number sets of social media messages files generated, number of output JSON files is: ", numFiles)
    return 


"""
    Main function
"""
if __name__ == '__main__':
    ##
    ##Read Reference Data from Spreadsheet files
    ##
    # Read Words to add to the start and end of the Twitter Message Text

    df_start_textlist = pd.read_excel("{}/{}".format(localRefDataFilePath,\
                                fileNamePhrases), sheet_name="start")
    df_end_textlist = pd.read_excel("{}/{}".format(localRefDataFilePath,\
                                fileNamePhrases), sheet_name="end")
    #
    # Read list of words to add to Twitter handle name
    nounListFile = open("{}/{}".format(localRefDataFilePath, randomNounsFileName), "r")
    nounList = nounListFile.read()
    nounList = nounList.split("\n")
    nounListFile.close()
    num_nouns=len(nounList)
    #
    # Read Baseline Message Generation Rates for each age
    base_reference_workbook= pd.read_excel("{}/{}".format(localRefDataFilePath,\
                                excelRefDataFileName), sheet_name=excelRefBaseWks)
    #
    # Read SNOMED Code Message Generation Rates
    code_reference_workbook = pd.read_excel("{}/{}".format(localRefDataFilePath, \
                                excelRefDataFileName), sheet_name=excelRefCodeWks)


    print('>> Read_FHIR_dates')
    # Read Data from the Synthea JSON files 
    data_dates = read_FHIR_dates()

    print('>> Obtain_twitter_dates')    
    # Identify the Patient's / User's that are the ages within the dates for using Twitter (TED)
    data_dates_twitter, filenames_unprocessed, filenames_processed = obtain_twitter_dates(data_dates)

    print('>> Config_stats_sheet')    
    # Create the message generation statistics spreadsheet
    config_stats_sheet(filenames_processed)

    print('>> Obtain_tweet_conditions')    
    # Add the relevant Medical Conditions from the Synthea JSON files
    data_dates_twitter = obtain_tweet_conditions(data_dates_twitter)

    startTime = time.time()
    print('>> Generate_tweets')    
    # Generate the social media messages (Tweets)
    generate_tweets(data_dates_twitter)
    endTime = time.time()
    # Print to the screen the number of files processed
    output_statistics(filenames_processed, filenames_unprocessed)

    print(endTime - startTime, " seconds...")
